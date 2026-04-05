from django.contrib import admin
from .models import CarouselItem, Service, DutyStaff, Course, Registration, SigninTemplate, Event, Story, USRAchievement, TechProject, ExperienceCourse, TechSection
import openpyxl
from django.http import HttpResponse
from django.utils.timezone import localtime

@admin.action(description='匯出所選報名資料為 Excel (.xlsx)')
def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="course_registrations.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '報名資料清單'
    
    columns = ['報名課程', '姓名', '性別', '聯絡電話', 'Email', '狀態', '人數', '報名時間']
    ws.append(columns)
    
    for obj in queryset:
        status = f'候補第 {obj.waitlist_number} 號' if obj.is_waitlisted else '正取'
        ws.append([
            obj.course.title,
            obj.name,
            obj.get_gender_display(),
            obj.phone,
            obj.email,
            status,
            obj.headcount,
            localtime(obj.created_at).strftime('%Y-%m-%d %H:%M:%S')
        ])
        
        wb.save(response)
    return response

@admin.action(description='批次匯出課程報名名單 (.xlsx)')
def export_course_registrations_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="course_registrations_export.xlsx"'
    
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for course in queryset:
        safe_title = str(course.title)[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_')
        ws = wb.create_sheet(title=safe_title)
        
        columns = ['報名課程', '姓名', '性別', '聯絡電話', 'Email', '狀態', '人數', '報名時間']
        ws.append(columns)
        
        registrations = Registration.objects.filter(course=course).order_by('created_at')
        for obj in registrations:
            status = f'候補第 {obj.waitlist_number} 號' if obj.is_waitlisted else '正取'
            ws.append([
                obj.course.title,
                obj.name,
                obj.get_gender_display(),
                obj.phone,
                obj.email,
                status,
                obj.headcount,
                localtime(obj.created_at).strftime('%Y-%m-%d %H:%M:%S')
            ])
            
    if not wb.sheetnames:
        wb.create_sheet(title='無資料')
        
    wb.save(response)
    return response

@admin.register(CarouselItem)
class CarouselItemAdmin(admin.ModelAdmin):
    list_display = ('title', 'order', 'show_title', 'show_subtitle', 'is_active')
    list_editable = ('order', 'show_title', 'show_subtitle', 'is_active')

@admin.register(Service)
class ServiceAdmin(admin.ModelAdmin):
    list_display = ('title', 'is_active')
    list_editable = ('is_active',)

@admin.register(DutyStaff)
class DutyStaffAdmin(admin.ModelAdmin):
    list_display = ('name', 'role', 'contact', 'is_on_duty')
    list_editable = ('is_on_duty',)
    list_filter = ('is_on_duty',)

@admin.action(description='📥 下載實體簽到表 (套版)')
def export_signin_sheet_to_word(modeladmin, request, queryset):
    from docxtpl import DocxTemplate
    from django.shortcuts import render
    from django.contrib.admin import helpers
    from django.utils.timezone import localtime
    from django.contrib import messages

    templates = SigninTemplate.objects.all()
    if not templates.exists():
        messages.warning(request, '系統中無任何可用範本，請先至「簽到表版型範本」上傳您的 Word 範本 (`.docx`)。')
        return None

    if 'apply' in request.POST:
        template_id = request.POST.get('template_id')
        selected_temp = SigninTemplate.objects.get(id=template_id)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = 'attachment; filename="course_signin_sheet.docx"'

        # 以第一堂課為主標題
        first_course = queryset.first()
        course_title = first_course.title
        course_date = localtime(first_course.start_time).strftime('%Y-%m-%d %H:%M')

        # 整理合併名單
        registrations_data = []
        global_idx = 1
        for course in queryset:
            regs = Registration.objects.filter(course=course, is_waitlisted=False).order_by('name')
            for reg in regs:
                registrations_data.append({
                    'index': global_idx,
                    'course': course.title,
                    'name': reg.name,
                    'gender': reg.get_gender_display(),
                    'phone': reg.phone,
                    'headcount': reg.headcount,
                })
                global_idx += 1
                
        context = {
            'course_title': course_title,
            'course_date': course_date,
            'registrations': registrations_data,
        }
        
        doc = DocxTemplate(selected_temp.file.path)
        doc.render(context)
        doc.save(response)
        return response

    context = {
        'title': "選擇簽到表範本",
        'queryset': queryset,
        'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
        'templates': templates,
        'opts': modeladmin.model._meta,
    }
    return render(request, 'admin/select_template.html', context)

@admin.register(SigninTemplate)
class SigninTemplateAdmin(admin.ModelAdmin):
    list_display = ('name', 'is_default', 'created_at')
    list_editable = ('is_default',)

@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    list_display = ('title', 'instructor', 'start_time', 'end_time', 'reg_start_time', 'reg_end_time', 'includes_lunch', 'capacity', 'is_active')
    list_editable = ('is_active', 'includes_lunch')
    list_filter = ('is_active', 'includes_lunch', 'start_time')
    actions = [export_course_registrations_to_excel, export_signin_sheet_to_word]

@admin.register(Registration)
class RegistrationAdmin(admin.ModelAdmin):
    list_display = ('name', 'course', 'phone', 'headcount', 'get_status', 'is_attended', 'created_at')
    list_filter = ('course', 'is_waitlisted', 'is_attended', 'created_at')
    search_fields = ('name', 'phone')
    actions = [export_to_excel]
    
    @admin.display(description='報名狀態')
    def get_status(self, obj):
        if obj.is_waitlisted:
            return f"候補第 {obj.waitlist_number} 號"
        return "正取"

@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = ('title', 'date')
    search_fields = ('title', 'description')
    list_filter = ('date',)

@admin.register(Story)
class StoryAdmin(admin.ModelAdmin):
    list_display = ('title', 'category', 'created_at')
    list_filter = ('category',)
    search_fields = ('title', 'content')

@admin.register(USRAchievement)
class USRAchievementAdmin(admin.ModelAdmin):
    list_display = ('title', 'type', 'date')
    list_filter = ('type', 'date')
    search_fields = ('title', 'summary')

class TechProjectInline(admin.StackedInline):
    model = TechProject
    extra = 1

@admin.register(TechSection)
class TechSectionAdmin(admin.ModelAdmin):
    list_display = ('title', 'layout_type', 'order', 'is_active')
    list_editable = ('order', 'is_active')
    inlines = [TechProjectInline]

@admin.register(TechProject)
class TechProjectAdmin(admin.ModelAdmin):
    list_display = ('name', 'section', 'order', 'is_active')
    list_editable = ('section', 'order', 'is_active')
    list_filter = ('section', 'is_active')

@admin.register(ExperienceCourse)
class ExperienceCourseAdmin(admin.ModelAdmin):
    list_display = ('name', 'category')
    list_filter = ('category',)
